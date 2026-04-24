from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, distinct, text
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Tuple
from pydantic import BaseModel, EmailStr
import pandas as pd
import io
import logging
import os
from pathlib import Path
from ..db.session import get_db

logger = logging.getLogger(__name__)
from ..models.machine_status import MachineCopyStatus
from ..models.metadata import ElImage
from ..models.path_config import PathConfig
from ..models.production_report import ProductionReport
from ..schemas.production_report import (
    ProductionSummaryRequest, 
    ProductionSummaryResponse, 
    ProductionReportResponse,
    ProductionChartResponse,
    ChartDataPoint
)
from ..routers.machineconfigure import EL_Machine, RFID_Machine, IVC_Machine
from ..core.security import get_current_admin_or_qualityadmin
from ..services.email_service import send_email_with_attachment

router = APIRouter(prefix="/api/reports", tags=["Reports"])

# Filesystem counting configuration (must match production folder conventions)
TYPE_DIR_MAP: Dict[str, str] = {
    "el": "EL-images",
    "rfid": "RFID",
    "ivc": "IVC",
}

# For non-Z machines: count only inside these folders.
# For Z machines: count only outside these folders.
NORMAL_FOLDERS = {"NORMAL", "OKE", "OK"}
RECURSIVE_MAX_DEPTH = 50  # Align with the standalone filesystem recursive scripts


def _is_z_machine(machine_name: str) -> bool:
    """
    Heuristic to detect 'Z' machines from their machine_name.
    Adjust if your naming convention differs.
    """
    if not machine_name:
        return False
    s = str(machine_name).upper()
    return s.startswith("Z") or "_Z_" in s or s.endswith("_Z")


def _get_machine_model(machine_type: str):
    mt = (machine_type or "").lower()
    if mt == "el":
        return EL_Machine
    if mt == "rfid":
        return RFID_Machine
    if mt == "ivc":
        return IVC_Machine
    return None


def _get_machine_dest_path(
    db: Session, machine_type: str, machine_name: str
) -> Optional[Path]:
    """
    Resolve destination base path from machine configure tables.
    Falls back to network_path if dest_path is not set.
    """
    model = _get_machine_model(machine_type)
    if not model:
        return None

    machine = (
        db.query(model)
        .filter(model.machine_name == machine_name)
        .first()
    )
    if not machine:
        return None

    raw_path = getattr(machine, "dest_path", None) or getattr(
        machine, "network_path", None
    )
    if not raw_path:
        return None

    try:
        return Path(raw_path)
    except TypeError:
        return None


def _iter_date_range(start_date: date, end_date: date):
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def _build_date_folder_candidates(base: Path, d: date) -> List[Path]:
    """
    Build possible date folder paths under a destination base.

    We support both zero-padded and non-padded month/day variants to match
    existing folder conventions, e.g. base / '2026' / '1' / '10'
    and base / '2026' / '01' / '10'.
    """
    year_str = f"{d.year}"
    month_str = f"{d.month}"
    month_str_padded = f"{d.month:02d}"
    day_str = f"{d.day}"
    day_str_padded = f"{d.day:02d}"

    candidates = [
        base / year_str / month_str / day_str,
        base / year_str / month_str / day_str_padded,
        base / year_str / month_str_padded / day_str,
        base / year_str / month_str_padded / day_str_padded,
    ]

    # If base already points to a year folder (endswith the year),
    # we also try without repeating the year segment.
    if base.name == year_str:
        candidates.extend(
            [
                base / month_str / day_str,
                base / month_str / day_str_padded,
                base / month_str_padded / day_str,
                base / month_str_padded / day_str_padded,
            ]
        )

    # Deduplicate while preserving order
    seen = set()
    unique_candidates: List[Path] = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            unique_candidates.append(c)
    return unique_candidates


def _build_machine_day_folder_candidates(
    base: Path,
    machine_type: str,
    machine_name: str,
    d: date,
) -> List[Path]:
    """
    Build possible day-folder paths for filesystem traversal.

    Production folder structure can be either:
      1) base / <year?> / <month> / <day>
      2) base / <TYPE_DIR> / <machine_name> / <month> / <day>
      3) base / <machine_name> / <month> / <day>

    We generate candidates for all these layouts and rely on `.is_dir()` checks.
    """
    year_str = f"{d.year}"
    month_str = f"{d.month}"
    month_str_padded = f"{d.month:02d}"
    day_str = f"{d.day}"
    day_str_padded = f"{d.day:02d}"

    type_dir = TYPE_DIR_MAP.get((machine_type or "").lower())

    # Preserve insertion order while deduplicating.
    seen = set()
    out: List[Path] = []

    def add(p: Path) -> None:
        if p not in seen:
            seen.add(p)
            out.append(p)

    # Layout 1 (existing): base / year / month / day  (plus base/month/day if base.name == year)
    for p in _build_date_folder_candidates(base, d):
        add(p)

    # Layout 1b: base / month / day (unconditionally; handles 'machine root' base)
    add(base / month_str / day_str)
    add(base / month_str / day_str_padded)
    add(base / month_str_padded / day_str)
    add(base / month_str_padded / day_str_padded)

    # Layout 3: base / machine_name / month / day
    if machine_name:
        mn_root = base / machine_name
        add(mn_root / month_str / day_str)
        add(mn_root / month_str / day_str_padded)
        add(mn_root / month_str_padded / day_str)
        add(mn_root / month_str_padded / day_str_padded)

        add(mn_root / year_str / month_str / day_str)
        add(mn_root / year_str / month_str / day_str_padded)
        add(mn_root / year_str / month_str_padded / day_str)
        add(mn_root / year_str / month_str_padded / day_str_padded)

    # Layout 2: base / TYPE_DIR / machine_name / month / day
    if type_dir and machine_name:
        typed_root = base / type_dir / machine_name
        add(typed_root / month_str / day_str)
        add(typed_root / month_str / day_str_padded)
        add(typed_root / month_str_padded / day_str)
        add(typed_root / month_str_padded / day_str_padded)

        add(typed_root / year_str / month_str / day_str)
        add(typed_root / year_str / month_str / day_str_padded)
        add(typed_root / year_str / month_str_padded / day_str)
        add(typed_root / year_str / month_str_padded / day_str_padded)

    return out


def _count_images_in_folder(folder: Path) -> int:
    """
    Count all files directly under the given folder, regardless of extension.
    We avoid deep recursive walks for performance.
    """
    if not folder.exists() or not folder.is_dir():
        return 0

    count = 0
    try:
        for entry in os.scandir(folder):
            if entry.is_file():
                count += 1
    except (OSError, PermissionError):
        # On any filesystem error we fall back to zero to avoid breaking reports
        return 0
    return count


def _count_images_recursive(
    folder: Path,
    *,
    is_z_machine: bool,
    max_depth: int,
    use_normal_policy: bool,
) -> int:
    """
    Recursive count of files under `folder`, with a max depth relative to `folder`.

    - If `use_normal_policy` is False: count all files.
    - If `use_normal_policy` is True:
        - Non-Z: count only when the relative path includes NORMAL/OKE/OK
        - Z: count only when the relative path does NOT include NORMAL/OKE/OK
    """
    if not folder.exists() or not folder.is_dir():
        return 0

    count = 0
    try:
        base = str(folder.resolve())
    except Exception:
        base = str(folder)

    # Walk and count files.
    for root, dirs, files in os.walk(str(folder)):
        # Depth relative to the base folder.
        try:
            rel = os.path.relpath(os.path.abspath(root), base)
            depth = 0 if rel == "." else rel.count(os.sep)
        except Exception:
            depth = 0

        if depth > max_depth:
            continue

        # Prune for Z machines: if we're already under NORMAL/OKE, deeper will also include it.
        if use_normal_policy and is_z_machine:
            try:
                rel_parts = [p.upper() for p in rel.split(os.sep) if p and p != "."]
            except Exception:
                rel_parts = []
            has_normal = any(p in NORMAL_FOLDERS for p in rel_parts)
            if has_normal:
                dirs[:] = []

        # Decide whether to count this subtree.
        include = True
        if use_normal_policy:
            try:
                rel_parts = [p.upper() for p in rel.split(os.sep) if p and p != "."]
            except Exception:
                rel_parts = []
            has_normal = any(p in NORMAL_FOLDERS for p in rel_parts)
            include = (not has_normal) if is_z_machine else has_normal

        if not include:
            continue

        for f in files:
            full = os.path.join(root, f)
            try:
                if os.path.isfile(full):
                    count += 1
            except Exception:
                pass

    return count


def _get_filesystem_image_count_for_machine_and_date(
    db: Session,
    machine_type: str,
    machine_name: str,
    d: date,
    dest_cache: Dict[Tuple[str, str], Optional[Path]],
    per_day_cache: Dict[Tuple[str, str, date], int],
) -> int:
    """
    Return image count for a given machine and date by inspecting
    the destination folder structure.
    """
    key = (machine_type, machine_name, d)
    if key in per_day_cache:
        return per_day_cache[key]

    dest_key = (machine_type, machine_name)
    if dest_key not in dest_cache:
        dest_cache[dest_key] = _get_machine_dest_path(db, machine_type, machine_name)

    base = dest_cache[dest_key]
    if not base:
        per_day_cache[key] = 0
        return 0

    is_z = _is_z_machine(machine_name)
    total = 0
    # Generate day folder candidates for the current machine & date.
    for folder in _build_machine_day_folder_candidates(
        base=base,
        machine_type=machine_type,
        machine_name=machine_name,
        d=d,
    ):
        # Use recursive traversal (policy-based), rather than short-circuiting
        # to "direct files only". This keeps chart/summary totals consistent with
        # the recursive counting logic you used previously.
        policy_count = _count_images_recursive(
            folder,
            is_z_machine=is_z,
            max_depth=RECURSIVE_MAX_DEPTH,
            use_normal_policy=True,
        )
        if policy_count > 0:
            total += policy_count
            continue

        # Final fallback: count all recursive files (avoid returning 0 due to policy mismatch).
        total += _count_images_recursive(
            folder,
            is_z_machine=is_z,
            max_depth=RECURSIVE_MAX_DEPTH,
            use_normal_policy=False,
        )

    per_day_cache[key] = total
    return total


def get_machine_name_by_id_and_type(db: Session, machine_id: int, machine_type: str):
    """Get machine name by ID and type"""
    if machine_type.lower() == 'el':
        machine = db.query(EL_Machine).filter(EL_Machine.id == machine_id).first()
    elif machine_type.lower() == 'rfid':
        machine = db.query(RFID_Machine).filter(RFID_Machine.id == machine_id).first()
    elif machine_type.lower() == 'ivc':
        machine = db.query(IVC_Machine).filter(IVC_Machine.id == machine_id).first()
    else:
        return f"Unknown_{machine_id}"
    
    return machine.machine_name if machine else f"{machine_type.upper()}_{machine_id}"

@router.get("/production-summary")
def get_production_summary(
    start_date: date = Query(..., description="Start date for report"),
    end_date: date = Query(..., description="End date for report"),
    machine_type: Optional[str] = Query(None, description="Filter by machine type (el, rfid, ivc)"),
    production_line: Optional[str] = Query(None, description="Filter by production line"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Get production summary for specified date range and filters"""
    
    # Convert dates to datetime for comparison
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # Get machine model based on type
    machine_model_map = {
        'el': EL_Machine,
        'rfid': RFID_Machine,
        'ivc': IVC_Machine
    }
    
    # Build query with proper joins to ensure only existing machines are included
    if machine_type:
        machine_model = machine_model_map.get(machine_type.lower())
        if not machine_model:
            return ProductionSummaryResponse(
                total_images_processed=0,
                total_valid_ids=0,
                total_failed_count=0,
                total_bytes_processed=0.0,
                success_rate=0.0,
                daily_reports=[],
                summary={
                    'total_days': 0,
                    'total_machines': 0,
                    'average_daily_processing': 0.0,
                    'average_success_rate': 0.0
                }
            )
        
        query = db.query(MachineCopyStatus).join(
            machine_model, MachineCopyStatus.machine_name == machine_model.machine_name
        ).filter(
            and_(
                MachineCopyStatus.cycle_start_time >= start_datetime,
                MachineCopyStatus.cycle_start_time <= end_datetime
            )
        )
        
        query = query.filter(MachineCopyStatus.machine_type == machine_type.lower())
        
        if production_line:
            query = query.filter(MachineCopyStatus.machine_name == production_line)
        
        records = query.all()
    else:
        # For all machine types, we need to handle each type separately
        all_records = []
        
        for mt, model in machine_model_map.items():
            mt_query = db.query(MachineCopyStatus).join(
                model, MachineCopyStatus.machine_name == model.machine_name
            ).filter(
                and_(
                    MachineCopyStatus.cycle_start_time >= start_datetime,
                    MachineCopyStatus.cycle_start_time <= end_datetime
                )
            )
            
            mt_query = mt_query.filter(MachineCopyStatus.machine_type == mt)
            
            if production_line:
                mt_query = mt_query.filter(MachineCopyStatus.machine_name == production_line)
            
            all_records.extend(mt_query.all())
        
        records = all_records
    
    # Calculate totals using filesystem-based counts per machine per day
    dest_cache: Dict[Tuple[str, str], Optional[Path]] = {}
    per_day_cache: Dict[Tuple[str, str, date], int] = {}
    total_images_processed = 0
    total_valid_ids = 0
    total_failed_count = sum(r.failed_count or 0 for r in records)
    total_bytes_processed = sum(r.bytes_copied or 0 for r in records)
    
    # Calculate success rate
    success_rate = (total_valid_ids / total_images_processed * 100) if total_images_processed > 0 else 0
    
    # Group by date and machine for detailed breakdown
    daily_reports = {}
    for record in records:
        if not record.cycle_start_time:
            continue
            
        report_date = record.cycle_start_time.date()
        key = f"{report_date}_{record.machine_type}_{record.machine_name}"
        
        if key not in daily_reports:
            daily_reports[key] = {
                'report_date': record.cycle_start_time.date(),
                'machine_type': record.machine_type,
                'production_line': record.machine_name,
                'total_failed_count': 0,
                'total_bytes_processed': 0.0,
                'processing_times': [],
                'image_count': None,
            }
        
        daily_reports[key]['total_failed_count'] += record.failed_count or 0
        daily_reports[key]['total_bytes_processed'] += record.bytes_copied or 0
        
        if record.copy_duration_seconds:
            daily_reports[key]['processing_times'].append(record.copy_duration_seconds)
    
    # For each (date, machine, type) group, compute filesystem image count once
    report_responses = []
    for key, data in daily_reports.items():
        d = data['report_date']
        mt = data['machine_type']
        line = data['production_line']
        images = _get_filesystem_image_count_for_machine_and_date(
            db=db,
            machine_type=mt,
            machine_name=line,
            d=d,
            dest_cache=dest_cache,
            per_day_cache=per_day_cache,
        )
        data['image_count'] = images
        total_images_processed += images
        total_valid_ids += images

        avg_time = sum(data['processing_times']) / len(data['processing_times']) if data['processing_times'] else 0
        
        report_responses.append(ProductionReportResponse(
            id=hash(key) % 1000000,  # Generate a pseudo ID
            report_date=datetime.combine(data['report_date'], datetime.min.time()),
            machine_type=data['machine_type'],
            production_line=data['production_line'],
            total_images_processed=data['image_count'] or 0,
            total_valid_ids=data['image_count'] or 0,
            total_failed_count=data['total_failed_count'],
            total_bytes_processed=data['total_bytes_processed'],
            average_processing_time=avg_time,
            created_at=datetime.now(),
            updated_at=datetime.now()
        ))
    
    return ProductionSummaryResponse(
        total_images_processed=total_images_processed,
        total_valid_ids=total_valid_ids,
        total_failed_count=total_failed_count,
        total_bytes_processed=total_bytes_processed,
        success_rate=round(success_rate, 2),
        reports=report_responses
    )

@router.get("/production-chart")
def get_production_chart_data(
    start_date: date = Query(..., description="Start date for chart"),
    end_date: date = Query(..., description="End date for chart"),
    machine_type: Optional[str] = Query(None, description="Filter by machine type"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Get production data formatted for charts"""
    
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # Build query
    query = db.query(MachineCopyStatus).filter(
        and_(
            MachineCopyStatus.cycle_start_time >= start_datetime,
            MachineCopyStatus.cycle_start_time <= end_datetime
        )
    )
    
    if machine_type:
        query = query.filter(MachineCopyStatus.machine_type == machine_type.lower())
    
    records = query.all()
    
    # Use filesystem-based counts but avoid double-counting:
    # If multiple DB rows exist for the same (date, machine_type, machine_name),
    # we must compute the filesystem count once for that unique tuple.
    unique_keys: set[Tuple[date, str, str]] = set()
    for record in records:
        if not record.cycle_start_time or not record.machine_type or not record.machine_name:
            continue
        unique_keys.add((record.cycle_start_time.date(), record.machine_type, record.machine_name))

    dest_cache: Dict[Tuple[str, str], Optional[Path]] = {}
    per_day_cache: Dict[Tuple[str, str, date], int] = {}
    chart_data: Dict[str, Dict[str, object]] = {}

    for d, machine_type_key, machine_name in unique_keys:
        date_str = d.strftime("%Y-%m-%d")

        images = _get_filesystem_image_count_for_machine_and_date(
            db=db,
            machine_type=machine_type_key,
            machine_name=machine_name,
            d=d,
            dest_cache=dest_cache,
            per_day_cache=per_day_cache,
        )

        key = f"{date_str}_{machine_type_key}"
        if key not in chart_data:
            chart_data[key] = {
                'date': date_str,
                'machine_type': machine_type_key,
                'total_images': 0
            }

        chart_data[key]['total_images'] = int(chart_data[key]['total_images']) + images
    
    # Convert to chart format
    chart_points = []
    for data in chart_data.values():
        chart_points.append(ChartDataPoint(
            date=data['date'],
            value=data['total_images'],
            machine_type=data['machine_type']
        ))
    
    # Sort by date
    chart_points.sort(key=lambda x: x.date)
    
    # Calculate summary
    total_images = sum(point.value for point in chart_points)
    machine_types = list(set(point.machine_type for point in chart_points))
    
    summary = {
        'total_images_processed': total_images,
        'date_range': f"{start_date} to {end_date}",
        'machine_types_included': machine_types,
        'total_data_points': len(chart_points)
    }
    
    return ProductionChartResponse(
        chart_data=chart_points,
        summary=summary
    )

@router.get("/machine-types")
def get_available_machine_types(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Get list of available machine types for filtering"""
    
    # Get machine types that have both copy status records AND existing machines
    machine_types = []
    
    # Check EL machines
    el_machines_with_status = db.query(distinct(MachineCopyStatus.machine_type)).join(
        EL_Machine, MachineCopyStatus.machine_name == EL_Machine.machine_name
    ).filter(MachineCopyStatus.machine_type == 'el').all()
    if el_machines_with_status:
        machine_types.extend([mt[0] for mt in el_machines_with_status if mt[0]])
    
    # Check RFID machines
    rfid_machines_with_status = db.query(distinct(MachineCopyStatus.machine_type)).join(
        RFID_Machine, MachineCopyStatus.machine_name == RFID_Machine.machine_name
    ).filter(MachineCopyStatus.machine_type == 'rfid').all()
    if rfid_machines_with_status:
        machine_types.extend([mt[0] for mt in rfid_machines_with_status if mt[0]])
    
    # Check IVC machines
    ivc_machines_with_status = db.query(distinct(MachineCopyStatus.machine_type)).join(
        IVC_Machine, MachineCopyStatus.machine_name == IVC_Machine.machine_name
    ).filter(MachineCopyStatus.machine_type == 'ivc').all()
    if ivc_machines_with_status:
        machine_types.extend([mt[0] for mt in ivc_machines_with_status if mt[0]])
    
    return list(set(machine_types))  # Remove duplicates

@router.get("/production-lines")
def get_available_production_lines(
    machine_type: Optional[str] = Query(None, description="Filter by machine type"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Get list of available production lines (machine names) for filtering"""
    
    # Get machine model based on type
    machine_model_map = {
        'el': EL_Machine,
        'rfid': RFID_Machine,
        'ivc': IVC_Machine
    }
    
    if machine_type:
        machine_model = machine_model_map.get(machine_type.lower())
        if not machine_model:
            return []
        
        # Join with machine table to ensure only existing machines are returned
        query = db.query(distinct(MachineCopyStatus.machine_name)).join(
            machine_model, MachineCopyStatus.machine_name == machine_model.machine_name
        ).filter(MachineCopyStatus.machine_type == machine_type.lower())
    else:
        # For all machine types, we need to check each type separately
        production_lines = []
        
        for mt, model in machine_model_map.items():
            lines = db.query(distinct(MachineCopyStatus.machine_name)).join(
                model, MachineCopyStatus.machine_name == model.machine_name
            ).filter(MachineCopyStatus.machine_type == mt).all()
            production_lines.extend([pl[0] for pl in lines if pl[0]])
        
        return list(set(production_lines))  # Remove duplicates
    
    production_lines = query.all()
    return [pl[0] for pl in production_lines if pl[0]]

@router.get("/latest-logs-per-machine")
def get_latest_logs_per_machine(
    start_date: Optional[date] = Query(None, description="Start date for filtering logs"),
    end_date: Optional[date] = Query(None, description="End date for filtering logs"),
    machine_type: Optional[str] = Query(None, description="Filter by machine type"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Get the latest log entry for each machine within the specified date range"""
    
    # Get machine model based on type
    machine_model_map = {
        'el': EL_Machine,
        'rfid': RFID_Machine,
        'ivc': IVC_Machine
    }
    
    # Build base query with date filters if provided, and join with machine table
    if machine_type:
        machine_model = machine_model_map.get(machine_type.lower())
        if not machine_model:
            return []
        
        base_query = db.query(MachineCopyStatus).join(
            machine_model, MachineCopyStatus.machine_name == machine_model.machine_name
        )
        
        if start_date and end_date:
            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(end_date, datetime.max.time())
            base_query = base_query.filter(
                and_(
                    MachineCopyStatus.cycle_start_time >= start_datetime,
                    MachineCopyStatus.cycle_start_time <= end_datetime
                )
            )
        
        base_query = base_query.filter(MachineCopyStatus.machine_type == machine_type.lower())
        
        # Get the latest cycle_end_time for each machine within the filtered data
        subquery = base_query.with_entities(
            MachineCopyStatus.machine_name,
            MachineCopyStatus.machine_type,
            func.max(MachineCopyStatus.cycle_end_time).label('latest_time')
        ).group_by(
            MachineCopyStatus.machine_name,
            MachineCopyStatus.machine_type
        ).subquery()
        
        # Join with main table to get full records
        query = db.query(MachineCopyStatus).join(
            subquery,
            and_(
                MachineCopyStatus.machine_name == subquery.c.machine_name,
                MachineCopyStatus.machine_type == subquery.c.machine_type,
                MachineCopyStatus.cycle_end_time == subquery.c.latest_time
            )
        )
        
        records = query.all()
    else:
        # For all machine types, we need to handle each type separately
        all_records = []
        
        for mt, model in machine_model_map.items():
            mt_query = db.query(MachineCopyStatus).join(
                model, MachineCopyStatus.machine_name == model.machine_name
            )
            
            if start_date and end_date:
                start_datetime = datetime.combine(start_date, datetime.min.time())
                end_datetime = datetime.combine(end_date, datetime.max.time())
                mt_query = mt_query.filter(
                    and_(
                        MachineCopyStatus.cycle_start_time >= start_datetime,
                        MachineCopyStatus.cycle_start_time <= end_datetime
                    )
                )
            
            mt_query = mt_query.filter(MachineCopyStatus.machine_type == mt)
            
            # Get the latest cycle_end_time for each machine within the filtered data
            subquery = mt_query.with_entities(
                MachineCopyStatus.machine_name,
                MachineCopyStatus.machine_type,
                func.max(MachineCopyStatus.cycle_end_time).label('latest_time')
            ).group_by(
                MachineCopyStatus.machine_name,
                MachineCopyStatus.machine_type
            ).subquery()
            
            # Join with main table to get full records
            query = db.query(MachineCopyStatus).join(
                subquery,
                and_(
                    MachineCopyStatus.machine_name == subquery.c.machine_name,
                    MachineCopyStatus.machine_type == subquery.c.machine_type,
                    MachineCopyStatus.cycle_end_time == subquery.c.latest_time
                )
            )
            
            all_records.extend(query.all())
        
        records = all_records
    
    # Convert to response format
    dest_cache: Dict[Tuple[str, str], Optional[Path]] = {}
    per_day_cache: Dict[Tuple[str, str, date], int] = {}

    latest_logs = []
    for record in records:
        # Prefer cycle_start_time date for filesystem counts (to match production-summary grouping).
        fs_day: Optional[date] = None
        if record.cycle_start_time:
            fs_day = record.cycle_start_time.date()
        elif record.cycle_end_time:
            fs_day = record.cycle_end_time.date()

        folder_valid_ids = 0
        if fs_day and record.machine_type and record.machine_name:
            folder_valid_ids = _get_filesystem_image_count_for_machine_and_date(
                db=db,
                machine_type=record.machine_type,
                machine_name=record.machine_name,
                d=fs_day,
                dest_cache=dest_cache,
                per_day_cache=per_day_cache,
            )

        latest_logs.append({
            'id': record.id,
            'machine_name': record.machine_name,
            'machine_type': record.machine_type,
            'cycle_start_time': record.cycle_start_time,
            'cycle_end_time': record.cycle_end_time,
            'status': record.status,
            # Make table totals folder-wise (filesystem traversal)
            'valid_ids_copied': folder_valid_ids,
            'failed_count': record.failed_count or 0,
            'total_ids_found': folder_valid_ids,
            'bytes_copied': record.bytes_copied or 0,
            'copy_duration_seconds': record.copy_duration_seconds,
            'error_message': record.error_message
        })
    
    return latest_logs

@router.get("/daily-summary")
def get_daily_production_summary(
    report_date: date = Query(..., description="Date for daily summary"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Get detailed daily production summary"""
    
    start_datetime = datetime.combine(report_date, datetime.min.time())
    end_datetime = datetime.combine(report_date, datetime.max.time())
    
    # Get all records for the day
    records = db.query(MachineCopyStatus).filter(
        and_(
            MachineCopyStatus.cycle_start_time >= start_datetime,
            MachineCopyStatus.cycle_start_time <= end_datetime
        )
    ).all()

    # Build summary shells from DB records (cycles, failures, bytes, active machines),
    # then compute pass totals from filesystem for consistency with charts/summary.
    summary_by_type: Dict[str, dict] = {}
    for record in records:
        mt = record.machine_type
        if not mt:
            continue

        if mt not in summary_by_type:
            summary_by_type[mt] = {
                'machine_type': mt,
                'total_images': 0,
                'total_valid': 0,
                'total_failed': 0,
                'total_bytes': 0,
                'machines': set(),
                'cycles': 0,
            }

        summary_by_type[mt]['total_failed'] += record.failed_count or 0
        summary_by_type[mt]['total_bytes'] += record.bytes_copied or 0
        summary_by_type[mt]['machines'].add(record.machine_name)
        summary_by_type[mt]['cycles'] += 1

    # Filesystem totals per distinct machine (computed once per day).
    dest_cache: Dict[Tuple[str, str], Optional[Path]] = {}
    per_day_cache: Dict[Tuple[str, str, date], int] = {}
    for mt, summary in summary_by_type.items():
        total_valid = 0
        for mn in summary['machines']:
            if not mn:
                continue
            total_valid += _get_filesystem_image_count_for_machine_and_date(
                db=db,
                machine_type=mt,
                machine_name=mn,
                d=report_date,
                dest_cache=dest_cache,
                per_day_cache=per_day_cache,
            )
        summary['total_images'] = total_valid
        summary['total_valid'] = total_valid
        summary['machines'] = list(summary['machines'])
        summary['success_rate'] = (100.0 if total_valid > 0 else 0)
    
    return {
        'date': report_date.isoformat(),
        'summary_by_type': list(summary_by_type.values()),
        'total_machines_active': sum(len(s['machines']) for s in summary_by_type.values()),
        'total_cycles': sum(s['cycles'] for s in summary_by_type.values()),
        'overall_total_images': sum(s['total_images'] for s in summary_by_type.values())
    }

@router.get("/machine-type-distribution")
def get_machine_type_distribution(
    start_date: date = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: date = Query(..., description="End date in YYYY-MM-DD format"),
    machine_type: Optional[str] = Query(None, description="Filter by specific machine type (el, rfid, ivc)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """
    Get machine type distribution by summing valid_ids for each machine type in the given date range.
    """
    try:
        # Convert dates to datetime for comparison
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())

        # Fetch candidate machines from DB (so we only compute for machines that were active),
        # but compute `total_valid_ids` from filesystem traversal.
        machines_query = db.query(
            MachineCopyStatus.machine_name,
            MachineCopyStatus.machine_type,
            func.max(MachineCopyStatus.cycle_end_time).label('last_copy_time')
        ).filter(
            and_(
                MachineCopyStatus.cycle_end_time >= start_datetime,
                MachineCopyStatus.cycle_end_time <= end_datetime,
                MachineCopyStatus.status == 'copied',
            )
        )

        if machine_type:
            machines_query = machines_query.filter(MachineCopyStatus.machine_type.ilike(f"%{machine_type}%"))

        machines_query = machines_query.group_by(
            MachineCopyStatus.machine_name,
            MachineCopyStatus.machine_type,
        )

        machines_result = machines_query.all()

        dest_cache: Dict[Tuple[str, str], Optional[Path]] = {}
        per_day_cache: Dict[Tuple[str, str, date], int] = {}

        machines_by_type: Dict[str, List[dict]] = {}
        total_valid_ids = 0

        for row in machines_result:
            mt = row.machine_type or ""
            mt_lower = mt.lower()
            mn = row.machine_name

            total_for_machine = 0
            for d in _iter_date_range(start_date, end_date):
                total_for_machine += _get_filesystem_image_count_for_machine_and_date(
                    db=db,
                    machine_type=mt,
                    machine_name=mn,
                    d=d,
                    dest_cache=dest_cache,
                    per_day_cache=per_day_cache,
                )

            total_valid_ids += total_for_machine
            machines_by_type.setdefault(mt_lower, []).append({
                "machine_name": mn,
                "total_valid_ids": total_for_machine,
                "last_copy_time": row.last_copy_time.isoformat() if row.last_copy_time else None
            })

        distribution = []
        for mt_lower, machines_list in machines_by_type.items():
            machines_list.sort(key=lambda x: x["total_valid_ids"], reverse=True)
            type_total = sum(m["total_valid_ids"] for m in machines_list)
            distribution.append({
                "machine_type": mt_lower,
                "total_valid_ids": type_total,
                "machines": machines_list,
            })

        distribution.sort(key=lambda x: x["total_valid_ids"], reverse=True)

        return {
            "success": True,
            "date_range": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat()
            },
            "total_valid_ids": total_valid_ids,
            "distribution": distribution
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch machine type distribution: {str(e)}")

class SendProductionReportRequest(BaseModel):
    email: EmailStr
    start_date: date
    end_date: date
    machine_type: Optional[str] = None
    production_line: Optional[str] = None

@router.post("/send-production-report")
def send_production_report_email(
    request: SendProductionReportRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_admin_or_qualityadmin)
):
    """Generate Excel production report and send it via email"""
    try:
        # Convert dates to datetime for comparison
        start_datetime = datetime.combine(request.start_date, datetime.min.time())
        end_datetime = datetime.combine(request.end_date, datetime.max.time())
        
        # Get machine model based on type
        machine_model_map = {
            'el': EL_Machine,
            'rfid': RFID_Machine,
            'ivc': IVC_Machine
        }
        
        # Build query with proper joins to ensure only existing machines are included
        if request.machine_type:
            machine_model = machine_model_map.get(request.machine_type.lower())
            if not machine_model:
                raise HTTPException(status_code=400, detail=f"Invalid machine type: {request.machine_type}")
            
            query = db.query(MachineCopyStatus).join(
                machine_model, MachineCopyStatus.machine_name == machine_model.machine_name
            ).filter(
                and_(
                    MachineCopyStatus.cycle_start_time >= start_datetime,
                    MachineCopyStatus.cycle_start_time <= end_datetime
                )
            )
            
            query = query.filter(MachineCopyStatus.machine_type == request.machine_type.lower())
            
            if request.production_line:
                query = query.filter(MachineCopyStatus.machine_name == request.production_line)
            
            records = query.all()
        else:
            # For all machine types, we need to handle each type separately
            all_records = []
            
            for mt, model in machine_model_map.items():
                mt_query = db.query(MachineCopyStatus).join(
                    model, MachineCopyStatus.machine_name == model.machine_name
                ).filter(
                    and_(
                        MachineCopyStatus.cycle_start_time >= start_datetime,
                        MachineCopyStatus.cycle_start_time <= end_datetime
                    )
                )
                
                mt_query = mt_query.filter(MachineCopyStatus.machine_type == mt)
                
                if request.production_line:
                    mt_query = mt_query.filter(MachineCopyStatus.machine_name == request.production_line)
                
                all_records.extend(mt_query.all())
            
            records = all_records
        
        # Prepare data for Excel: only line (machine_name) and total IDs generated
        # We'll generate the Excel with openpyxl to support styling.

        # Compute pass totals from filesystem so Excel matches charts + table.
        dest_cache: Dict[Tuple[str, str], Optional[Path]] = {}
        per_day_cache: Dict[Tuple[str, str, date], int] = {}

        date_days = list(_iter_date_range(request.start_date, request.end_date))

        unique_machine_keys: set[Tuple[str, str]] = set()
        for r in records:
            if r.machine_type and r.machine_name:
                unique_machine_keys.add((r.machine_type, r.machine_name))

        # One production line per machine_type, with totals in the selected date range.
        # Example output layout:
        #   Header row / Date row
        #   EL section (colored)
        #     Production Line | Total IDs Generated
        #   RFID section (colored)
        #     Production Line | Total IDs Generated
        #   IVC section (colored)
        #     Production Line | Total IDs Generated
        per_type_lines: Dict[str, Dict[str, int]] = {}
        for mt, line_name in unique_machine_keys:
            total_for_line = 0
            for d in date_days:
                total_for_line += _get_filesystem_image_count_for_machine_and_date(
                    db=db,
                    machine_type=mt,
                    machine_name=line_name,
                    d=d,
                    dest_cache=dest_cache,
                    per_day_cache=per_day_cache,
                )
            mt_lower = (mt or "").lower()
            per_type_lines.setdefault(mt_lower, {})[line_name] = total_for_line

        preferred_order = ["el", "rfid", "ivc"]
        type_order = [t for t in preferred_order if t in per_type_lines] + [
            t for t in sorted(per_type_lines.keys()) if t not in preferred_order
        ]

        # Azure upload summary (IST day-wise), based on metadata rows that were pushed to cloud.
        azure_params = {
            "start_date": request.start_date.isoformat(),
            "end_date": request.end_date.isoformat(),
        }
        azure_type_filter_sql = ""
        if request.machine_type:
            azure_params["machine_type"] = request.machine_type.lower()
            azure_type_filter_sql = " AND machine_type = :machine_type"

        azure_sql = text(f"""
            WITH typed AS (
                SELECT
                    DATE((created_at AT TIME ZONE 'UTC') AT TIME ZONE 'Asia/Kolkata') AS day,
                    CASE
                        WHEN lower(file_name) ~ '\\.(jpg|jpeg|png|gif|bmp|tiff|webp)$' THEN 'el'
                        WHEN lower(file_name) ~ '\\.pdf$' THEN 'rfid'
                        WHEN lower(file_name) ~ '\\.ivc$' THEN 'ivc'
                        ELSE 'other'
                    END AS machine_type
                FROM metadata.elimages
                WHERE DATE((created_at AT TIME ZONE 'UTC') AT TIME ZONE 'Asia/Kolkata') >= :start_date
                  AND DATE((created_at AT TIME ZONE 'UTC') AT TIME ZONE 'Asia/Kolkata') <= :end_date
                  AND azure_path IS NOT NULL
                  AND trim(azure_path) <> ''
            )
            SELECT
                day,
                machine_type,
                COUNT(*) AS file_count
            FROM typed
            WHERE machine_type IN ('el', 'rfid', 'ivc')
            {azure_type_filter_sql}
            GROUP BY day, machine_type
            ORDER BY day ASC, machine_type ASC;
        """)
        azure_daily_rows = db.execute(azure_sql, azure_params).mappings().all()
        azure_total_files = sum(int(r["file_count"]) for r in azure_daily_rows)

        # Active paths for report context.
        active_paths_query = db.query(PathConfig).filter(PathConfig.is_active == True)
        if request.machine_type:
            mt = request.machine_type.lower()
            active_paths_query = active_paths_query.filter(PathConfig.machine_type.in_([mt, "all"]))
        else:
            active_paths_query = active_paths_query.filter(
                PathConfig.machine_type.in_(["el", "rfid", "ivc", "all"])
            )
        active_paths = active_paths_query.order_by(PathConfig.machine_type.asc(), PathConfig.id.asc()).all()

        # Machine type colors (hex without '#')
        type_colors: Dict[str, Dict[str, str]] = {
            "el": {"fill": "4F81BD", "font": "1F4E79"},    # Blue
            "rfid": {"fill": "F79646", "font": "C55A11"},  # Orange
            "ivc": {"fill": "9BBB59", "font": "385723"},   # Green
        }

        default_fill = "D9D9D9"
        default_font = "000000"

        # Create styled Excel file in memory
        try:
            # Import here to avoid crashing API startup if openpyxl isn't installed.
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ModuleNotFoundError as e:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is required to generate the styled production report Excel.",
            ) from e

        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Production Report"

        thin_side = Side(style="thin", color="D9D9D9")
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Title
        title = "Tata Power Renewable Energy Ltd , Bangalore"
        row = 1
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = worksheet.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Date + subtitle
        row += 1
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = worksheet.cell(row=row, column=1, value="Production Report")
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")

        row += 1
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = worksheet.cell(
            row=row,
            column=1,
            value=f"Date Range: {request.start_date} to {request.end_date}",
        )
        c.font = Font(bold=False, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Spacing
        row += 2

        for mt_lower in type_order:
            type_cfg = type_colors.get(mt_lower, {"fill": default_fill, "font": default_font})
            fill_color = type_cfg["fill"]
            font_color = type_cfg["font"]

            # Section header
            row += 0
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            sec = worksheet.cell(row=row, column=1, value=f"{mt_lower.upper()} Machine Type")
            sec.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            sec.font = Font(bold=True, color="FFFFFF")
            sec.alignment = Alignment(horizontal="left", vertical="center")
            sec.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            row += 1

            # Table header row
            header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            ws_h1 = worksheet.cell(row=row, column=1, value="Production Line")
            ws_h2 = worksheet.cell(row=row, column=2, value="Total IDs Generated")
            for cell in (ws_h1, ws_h2):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = table_border

            row += 1

            # Data rows
            data_rows = per_type_lines.get(mt_lower, {})
            for line_name, total in sorted(data_rows.items(), key=lambda x: (str(x[0]).lower())):
                c1 = worksheet.cell(row=row, column=1, value=line_name)
                c2 = worksheet.cell(row=row, column=2, value=int(total or 0))

                c1.font = Font(color=font_color)
                c1.alignment = Alignment(horizontal="left", vertical="center")
                c1.border = table_border

                c2.number_format = "#,##0"
                c2.alignment = Alignment(horizontal="right", vertical="center")
                c2.border = table_border

                row += 1

            # Empty row between machine types
            row += 1

        # Column widths + freeze panes for readability
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 20
        # Freeze after the date rows; keeps section/table headers visible when scrolling
        worksheet.freeze_panes = "A6"

        # Second sheet: Azure upload details + active paths.
        azure_ws = workbook.create_sheet(title="Azure Upload Summary")
        azure_ws.column_dimensions["A"].width = 15
        azure_ws.column_dimensions["B"].width = 18
        azure_ws.column_dimensions["C"].width = 24
        azure_ws.column_dimensions["D"].width = 18
        azure_ws.column_dimensions["E"].width = 22

        row2 = 1
        azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
        c = azure_ws.cell(row=row2, column=1, value="Azure Upload Summary")
        c.font = Font(bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")

        row2 += 1
        azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
        c = azure_ws.cell(
            row=row2,
            column=1,
            value=f"Date Range (IST): {request.start_date} to {request.end_date}",
        )
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

        row2 += 1
        azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
        c = azure_ws.cell(
            row=row2,
            column=1,
            value=f"Total Files Pushed to Cloud: {azure_total_files:,}",
        )
        c.font = Font(bold=True, size=11, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")

        row2 += 2
        sec = azure_ws.cell(row=row2, column=1, value="Active Paths")
        sec.font = Font(bold=True, color="FFFFFF")
        sec.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        sec.alignment = Alignment(horizontal="left", vertical="center")
        sec.border = table_border
        azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)

        row2 += 1
        active_headers = ["Machine Type", "Active Path", "Azure Folder", "Description", "Status"]
        for idx, label in enumerate(active_headers, start=1):
            h = azure_ws.cell(row=row2, column=idx, value=label)
            h.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            h.font = Font(bold=True)
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = table_border

        row2 += 1
        if active_paths:
            for p in active_paths:
                values = [
                    (p.machine_type or "").upper(),
                    p.path or "",
                    p.azure_folder_name or "-",
                    p.description or "-",
                    "Active" if p.is_active else "Inactive",
                ]
                for idx, val in enumerate(values, start=1):
                    cell = azure_ws.cell(row=row2, column=idx, value=val)
                    cell.alignment = Alignment(
                        horizontal="left" if idx in (2, 3, 4) else "center",
                        vertical="center",
                    )
                    cell.border = table_border
                row2 += 1
        else:
            azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
            cell = azure_ws.cell(row=row2, column=1, value="No active paths configured.")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border
            row2 += 1

        row2 += 1
        sec2 = azure_ws.cell(row=row2, column=1, value="Daily Cloud Upload Details")
        sec2.font = Font(bold=True, color="FFFFFF")
        sec2.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        sec2.alignment = Alignment(horizontal="left", vertical="center")
        sec2.border = table_border
        azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)

        row2 += 1
        azure_headers = ["Day (IST)", "Machine Type", "Files Pushed", "Remarks"]
        for idx, label in enumerate(azure_headers, start=1):
            h = azure_ws.cell(row=row2, column=idx, value=label)
            h.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            h.font = Font(bold=True)
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = table_border

        row2 += 1
        if azure_daily_rows:
            for item in azure_daily_rows:
                pushed = int(item["file_count"] or 0)
                row_values = [
                    str(item["day"]),
                    (item["machine_type"] or "").upper(),
                    pushed,
                    "Uploaded to Azure",
                ]
                for idx, val in enumerate(row_values, start=1):
                    cell = azure_ws.cell(row=row2, column=idx, value=val)
                    if idx == 3:
                        cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = table_border
                row2 += 1
        else:
            azure_ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
            cell = azure_ws.cell(row=row2, column=1, value="No Azure uploads found for selected filters.")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border
            row2 += 1

        azure_ws.freeze_panes = "A7"

        output.seek(0)
        workbook.save(output)
        output.seek(0)
        excel_bytes = output.read()
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        machine_type_str = request.machine_type.upper() if request.machine_type else "ALL"
        filename = f"Production_Report_{machine_type_str}_{request.start_date}_{request.end_date}_{timestamp}.xlsx"
        
        # Email body: IVC-only, plain-text aligned columns (no borders).
        ivc_rows: List[Tuple[str, int]] = [
            (str(line_name), int(line_total or 0))
            for line_name, line_total in sorted(
                per_type_lines.get("ivc", {}).items(),
                key=lambda x: (str(x[0]).lower()),
            )
        ]

        if ivc_rows:
            w_line = max(len("Production line"), max(len(r[0]) for r in ivc_rows))
            w_ids = max(len("IDs generated"), max(len(f"{r[1]:,}") for r in ivc_rows))
            sep = f"{'-' * w_line}  {'-' * w_ids}"
            header = f"{'Production line'.ljust(w_line)}  {'IDs generated'.rjust(w_ids)}"
            data_lines = [
                f"{line_name.ljust(w_line)}  {ids:>{w_ids},}"
                for line_name, ids in ivc_rows
            ]
            ids_table = "\n".join([header, sep, *data_lines])
        else:
            ids_table = "No IVC production lines with IDs in the selected date range."

        email_body = f"""
Production Report
Date Range: {request.start_date} to {request.end_date}

IDs generated by production line:
{ids_table}

Finalized report attached as Excel.
"""
        
        # Send email with attachment
        email_sent = send_email_with_attachment(
            email=request.email,
            subject=f"Production Report - {request.start_date} to {request.end_date}",
            body=email_body,
            attachment_data=excel_bytes,
            attachment_filename=filename
        )
        
        if email_sent:
            return {
                "success": True,
                "message": f"Production report sent successfully to {request.email}",
                "filename": filename
            }
        else:
            raise HTTPException(
                status_code=500,
                detail="Failed to send email. Please check the email service configuration."
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating and sending production report: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate and send production report: {str(e)}")
